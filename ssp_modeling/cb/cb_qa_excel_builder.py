import pandas as pd 
from openpyxl import Workbook
from typing import Dict, List
import string 
import numpy as np 

class QAExcelBuilder:
    """
    Clase que genera un excel para hacer el QA de costos y beneficios
    """
    def __init__(self, 
                time_init : int,
                time_end : int,
                mapping_cb_diff_var : Dict[str, str], # Diccionario que mapea la variable de costo con la variable de sisepuede que utiliza para su cálculo
                #mapping_cb_benefits_aggregated : Dict[str, List[str]], # Diccionario que mapea la categía agregada de costos con las variables de costo correspondientes
                cb_cost_factors : pd.DataFrame, # Dataframe con los factores de costo
                contador_linea : int = 2,
                ):

        self.wb = Workbook()
        self.time_init = time_init
        self.time_end = time_end
        self.contador_linea = contador_linea
        self.contador_linea_backup = contador_linea
        self.mapping_cbvar_position = {} # Dictionary que mapea la variable de costo con su posición en la pestaña
        self.mapping_cb_value_position = {} # Diccionario que mapea la variable de costo con su valor calculado
        self.mapping_cbvar_aggregate_cumulative_position = {} 
        self.mapping_cb_diff_var = mapping_cb_diff_var
        self.cb_cost_factors = cb_cost_factors
        self.column_identifier_excel = self.crea_letras_excel()

    ########################################
    #------ METODOS DE INICIALIZACION------#
    ########################################
    
    def crea_letras_excel(
        self
        ) -> List[str]:

        letras_excel = list(string.ascii_uppercase)
        letras_excel += [f"A{i}"for i in string.ascii_uppercase]
        letras_excel += [f"B{i}"for i in string.ascii_uppercase]

        return letras_excel
    

    ########################################
    #------------- UTILITIES   ------------#
    ########################################

    def build_cb_value_formula(
        self,
        cb_var_name : str, 
        column_index : str,
        exponencial : float,
        df_base : pd.DataFrame
        ) -> str:
        

        if self.mapping_cb_diff_var[cb_var_name]:
            
            acumula_formulas = []

            exponencial = np.maximum(0, exponencial - 2)

            for ssp_var_name in self.mapping_cb_diff_var[cb_var_name]:
                baseline_diff_var_position = self.mapping_cbvar_position[ssp_var_name]
                tx_diff_var_position = self.mapping_cbvar_position[ssp_var_name] + df_base.shape[1]  + 2

                cb_factor_position = self.mapping_cbvar_position[cb_var_name]

                acumula_formulas.append(
                    f"(({column_index}{tx_diff_var_position}-{column_index}{baseline_diff_var_position})*$B${cb_factor_position}*$D${cb_factor_position}^{exponencial})"
                )

            return "=" +  "(" +"+".join(acumula_formulas) + ")/1e9" 
        else:
            return "=0.0"

    def activa_sheet(
        self,
        strategy_name : str
        ) -> None:

        self.wb.create_sheet(strategy_name)
        self.wb.active = self.wb.sheetnames.index(strategy_name)
        self.ws = self.wb.active

    #########################################
    #--------------- METHODS	   ---------#
    #########################################

    def set_mapping_cb_benefits_aggregated(self, 
                                           mapping_cb_benefits_aggregated : Dict[str, List[str]]) -> None:
            self.mapping_cb_benefits_aggregated = mapping_cb_benefits_aggregated

    def agrega_columnas_anios(
        self
        ) -> None:
        
        self.ws.append([""] + list(range(self.time_init, self.time_end + 1)))

    def add_baseline_data(
        self,
        data_baseline : pd.DataFrame
        ) -> None:

        self.ws[f'A{self.contador_linea}'] = "Baseline"
        self.mapping_cbvar_position["Baseline"] = self.contador_linea

        for row in data_baseline.T.reset_index().values.tolist():
            self.ws.append(row)

        self.mapping_cbvar_position.update({j:i+self.contador_linea+1 for i,j in enumerate(data_baseline.columns)})

        self.contador_linea += data_baseline.shape[1] + 2


    def add_pathway_data(
        self,
        data_pathway : pd.DataFrame
        ) -> None:

        ## Agrega datos del patwhay
        self.ws[f'A{self.contador_linea}'] = "Pathway"

        for row in data_pathway.T.reset_index().values.tolist():
            self.ws.append(row)

        self.contador_linea += data_pathway.shape[1] + 2

    def add_cost_factors(
        self
        ) -> None:

        self.ws[f'A{self.contador_linea}'] = "Cost Factors"

        for row in self.cb_cost_factors.T.reset_index().values.tolist():
            self.ws.append(row)

        self.mapping_cbvar_position.update({j:i+self.contador_linea+1 for i,j in enumerate(self.cb_cost_factors.columns)})

        self.contador_linea += self.cb_cost_factors.shape[1] + 2

    def compute_cost_values(
        self,
        df_base : pd.DataFrame
        ) -> None:

        self.ws[f'A{self.contador_linea}'] = "Cost Variables"
        self.contador_linea += 1

        for cb_var_name_iter in self.mapping_cb_diff_var:

            self.mapping_cb_value_position[cb_var_name_iter] = self.contador_linea

            self.ws.append(
                [cb_var_name_iter] + [
                    self.build_cb_value_formula(cb_var_name = cb_var_name_iter, 
                                column_index = self.column_identifier_excel[i],
                                exponencial = i,
                                df_base = df_base)

                    for i in range(1, self.time_end - self.time_init + 2)
                        
                ]
            )    

            self.contador_linea += 1    
        
        self.contador_linea += 2    

    def compute_aggregated_categories(
        self,
        ) -> None:

        self.ws[f'A{self.contador_linea}'] = "Cost-Benefit Aggregated"


        for cb_aggregate_val, cb_vals_to_agg in self.mapping_cb_benefits_aggregated.items():
            self.contador_linea += 1

            if cb_vals_to_agg:
                costos_beneficios_tiempo = [cb_aggregate_val] + ["="+"+".join([f"{self.column_identifier_excel[period]}{self.mapping_cb_value_position[i]}" for i in cb_vals_to_agg if i in self.mapping_cb_value_position])
                                    for period in range(1, self.time_end - self.time_init + 2)]
            else:
                costos_beneficios_tiempo = [cb_aggregate_val] + ["=0" ]

            self.ws.append(
                costos_beneficios_tiempo
            )

            self.mapping_cbvar_position[cb_aggregate_val] = self.contador_linea


        self.contador_linea += 2


    def compute_aggregated_cumulative_categories(
        self,
        ) -> None:

        self.ws[f'A{self.contador_linea}'] = "Cost-Benefit Cumulative"

                
        for cb_aggregate_val in self.mapping_cb_benefits_aggregated.keys():
            
            posiciones_beneficios = [f"{self.column_identifier_excel[period]}{self.mapping_cbvar_position[cb_aggregate_val]}" 
            for period in range(1, self.time_end - self.time_init + 2) ]

            if posiciones_beneficios:
                beneficios_acumulados_formula = [cb_aggregate_val] + ["="+ "+".join(posiciones_beneficios)]
            else:
                beneficios_acumulados_formula = [cb_aggregate_val] + ["=0"]

            self.ws.append(beneficios_acumulados_formula)
            self.contador_linea += 1 
            self.mapping_cbvar_aggregate_cumulative_position[cb_aggregate_val] = self.contador_linea

        
        self.contador_linea += 2    

    def add_test_cumulative_categories(
        self,
        df_cumulative_categories : pd.DataFrame,
        ) -> None:

        self.ws[f'A{self.contador_linea}'] = "Cost-Benefit Cumulative Test"

        for row in df_cumulative_categories.T.reset_index().values.tolist():
            self.ws.append(row)

        self.contador_linea += df_cumulative_categories.shape[1] + 2


    def compute_marginal_effects(
        self,
        emission_co2e_total_diff : float,
        ) -> None:

        self.ws[f'A{self.contador_linea}'] = "Cost-Benefit Marginal Effects"
        self.contador_linea += 1 

        ## Agrega Emission CO2E Total Diff
        self.ws.append(
            ["Emission CO2E Total Diff", emission_co2e_total_diff]
        )

        self.mapping_cbvar_position["emission_co2e_total_diff"] = self.contador_linea

        for cb_aggregate_val in self.mapping_cb_benefits_aggregated.keys():

            self.ws.append(
            [cb_aggregate_val] + ["=" + f"(B{self.mapping_cbvar_aggregate_cumulative_position[cb_aggregate_val]}/ABS(B{self.mapping_cbvar_position['emission_co2e_total_diff']}))*1000"]
            )

            self.contador_linea += 1 
        
        self.contador_linea += 2    


    def reset_contadores(
        self
        ) -> None:

        self.mapping_cbvar_position = {} 
        self.mapping_cb_value_position = {} 
        self.contador_linea = self.contador_linea_backup

    def compute_strategy_sheet(
        self,
        strategy_name : str,
        data_baseline : pd.DataFrame,
        data_pathway : pd.DataFrame,
        df_cumulative_categories : pd.DataFrame,
        emission_co2e_total_diff : float
        ) -> None:

        self.activa_sheet(strategy_name)
        self.agrega_columnas_anios()
        self.add_baseline_data(data_baseline)
        self.add_pathway_data(data_pathway)
        self.add_cost_factors()
        self.compute_cost_values(df_base = data_baseline)
        self.compute_aggregated_categories()
        self.compute_aggregated_cumulative_categories()
        self.add_test_cumulative_categories(df_cumulative_categories)
        self.compute_marginal_effects(emission_co2e_total_diff)
        self.reset_contadores()